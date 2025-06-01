import csv
import os
import time
import json
import base64
from openai import OpenAI
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Load environment variables
load_dotenv()
client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))

# Configuration
MODELS = [
    "gpt-4o-mini",
    "gpt-4o",
    "gpt-4.1",
    "gpt-4.1-mini",
    "gpt-4.1-nano",
    "o1",
    "o1-mini",
    "o3",
    "o3-mini",
    "o4-mini",
]

class Benchmark:
    def __init__(self, tasks_csv="data/tasks.csv"):
        self.tasks_csv = tasks_csv
        self.results = []
        self.tasks = self.load_tasks()
    
    def load_tasks(self):
        """Load tasks from CSV file"""
        tasks = []
        with open(self.tasks_csv, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=";")
            for row in reader:
                task = {
                    "question": row["question"].strip(),
                    "options": [x.strip() for x in row["options"].split(";")],
                    "correct_solution": row["correct_solution"].strip(),
                    "image_path": row.get("image_path", "").strip()
                }
                if task["correct_solution"] not in task["options"]:
                    task["options"].append(task["correct_solution"])
                tasks.append(task)
        return tasks
    
    def run_task(self, model, task):
        """Run a single task with the specified model"""
        start_time = time.time()
        
        # Prepare text for prompt
        prompt_text = f"Solve this math problem. Only provide the answer, no explanation or working: {task['question']}\n"
        prompt_text += f"Available options: {', '.join(task['options'])}\n"
        prompt_text += "Answer with just the correct option text."
        
        # Create content list starting with text
        content = [{"type": "text", "text": prompt_text}]
        
        # Add image if available
        if task.get("image_path"):
            img_path = os.path.join("static", "img", task["image_path"] + ".jpg")
            if os.path.exists(img_path):
                with open(img_path, "rb") as img_file:
                    # Convert image to base64
                    base64_image = base64.b64encode(img_file.read()).decode('utf-8')
                    
                    # Add image to content
                    content.append({
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    })
        
        try:
            # Using the responses.create method for model evaluation
            if model in ["gpt-4o-mini", "gpt-4o", "gpt-4.1", "gpt-4.1-mini", "gpt-4.1-nano"]:
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "You are a math problem solver. Provide only the answer, no explanation."},
                        {"role": "user", "content": content}
                    ],
                    max_tokens=50,
                    temperature=0
                )
            else:
                response = client.chat.completions.create(
                    model=model,
                    messages=[
                        {"role": "system", "content": "You are a math problem solver. Provide only the answer, no explanation."},
                        {"role": "user", "content": content}
                    ],
                )
            
            answer = response.choices[0].message.content.strip()
            elapsed_time = time.time() - start_time
            
            # Check if answer is correct
            is_correct = answer == task["correct_solution"] or task["correct_solution"] in answer
            
            return {
                "model": model,
                "question": task["question"],
                "correct_answer": task["correct_solution"],
                "model_answer": answer,
                "is_correct": is_correct,
                "time": elapsed_time
            }
        except Exception as e:
            print(e)
            return {
                "model": model,
                "question": task["question"],
                "correct_answer": task["correct_solution"],
                "model_answer": f"ERROR: {str(e)}",
                "is_correct": False,
                "time": time.time() - start_time
            }
    
    def run_benchmark(self, max_workers=5):
        """Run benchmark for all models and all tasks"""
        results = []
        
        for model in MODELS:
            model_start_time = time.time()
            model_results = []
            
            print(f"Testing model: {model}")
            
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                # Submit all tasks for this model
                future_to_task = {
                    executor.submit(self.run_task, model, task): task 
                    for task in self.tasks
                }
                
                # Process results as they complete
                for i, future in enumerate(as_completed(future_to_task)):
                    task = future_to_task[future]
                    try:
                        result = future.result()
                        model_results.append(result)
                        print(f"  - Task {i+1}/{len(self.tasks)} completed: {'✓' if result['is_correct'] else '✗'}")
                    except Exception as e:
                        print(f"  - Task {i+1}/{len(self.tasks)} failed: {str(e)}")
            
            # Calculate model statistics
            correct_count = sum(1 for r in model_results if r["is_correct"])
            total_time = sum(r["time"] for r in model_results)
            model_time = time.time() - model_start_time
            
            model_summary = {
                "model": model,
                "accuracy": correct_count / len(self.tasks) if self.tasks else 0,
                "correct_count": correct_count,
                "total_tasks": len(self.tasks),
                "avg_task_time": total_time / len(self.tasks) if self.tasks else 0,
                "total_time": total_time,
                "wall_clock_time": model_time,
                "results": model_results
            }
            
            results.append(model_summary)
            
            print(f"\nModel {model} summary:")
            print(f"  - Accuracy: {model_summary['accuracy']:.2%} ({model_summary['correct_count']}/{model_summary['total_tasks']})")
            print(f"  - Avg task time: {model_summary['avg_task_time']:.2f}s")
            print(f"  - Total processing time: {model_summary['total_time']:.2f}s")
            print(f"  - Wall clock time: {model_summary['wall_clock_time']:.2f}s\n")
        
        self.results = results
        return results
    
    def save_results(self, filename="benchmark_results.json", excel_filename="benchmark_results.xlsx"):
        """Save benchmark results to a file"""
        # Save to JSON
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(self.results, f, indent=2)
        print(f"Results saved to {filename}")
        
        # Save to Excel
        self.save_results_to_excel(excel_filename)
        print(f"Results also saved to {excel_filename}")
    
    def save_results_to_excel(self, filename="benchmark_results.xlsx"):
        """Save benchmark results to an Excel file with formatting"""
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Set up styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        # Create summary sheet
        ws_summary.append(["Model", "Accuracy", "Correct", "Total Tasks", "Avg Time (s)", "Total Time (s)"])
        
        # Apply header styles
        for cell in ws_summary[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Sort results by accuracy
        sorted_results = sorted(self.results, key=lambda x: x["accuracy"], reverse=True)
        
        # Add data rows
        for result in sorted_results:
            ws_summary.append([
                result["model"],
                f"{result['accuracy']:.2%}",
                result["correct_count"],
                result["total_tasks"],
                f"{result['avg_task_time']:.2f}",
                f"{result['total_time']:.2f}"
            ])
        
        # Auto-adjust column widths
        for col in ws_summary.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            ws_summary.column_dimensions[column].width = adjusted_width
        
        # Create detailed sheets for each model
        for result in self.results:
            model_name = result["model"]
            # Create a valid sheet name (max 31 chars, no invalid chars)
            sheet_name = model_name[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_')
            ws_detail = wb.create_sheet(title=sheet_name)
            
            # Add headers
            ws_detail.append(["Question", "Correct Answer", "Model Answer", "Correct?", "Time (s)"])
            for cell in ws_detail[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Add task results
            for task_result in result["results"]:
                ws_detail.append([
                    task_result["question"],
                    task_result["correct_answer"],
                    task_result["model_answer"],
                    "✓" if task_result["is_correct"] else "✗",
                    f"{task_result['time']:.2f}"
                ])
            
            # Auto-adjust column widths
            for col in ws_detail.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, min(100, len(str(cell.value))))
                adjusted_width = max_length + 2
                ws_detail.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        wb.save(filename)

if __name__ == "__main__":
    print("Starting benchmark of OpenAI models on math tasks...")
    benchmark = Benchmark()
    results = benchmark.run_benchmark()
    
    # Print final summary
    print("\n=== FINAL BENCHMARK RESULTS ===")
    print(f"Total tasks: {len(benchmark.tasks)}")
    
    # Sort models by accuracy
    sorted_results = sorted(results, key=lambda x: x["accuracy"], reverse=True)
    
    for i, result in enumerate(sorted_results):
        print(f"\n{i+1}. {result['model']}:")
        print(f"   Accuracy: {result['accuracy']:.2%} ({result['correct_count']}/{result['total_tasks']})")
        print(f"   Average time per task: {result['avg_task_time']:.2f}s")
        print(f"   Total processing time: {result['total_time']:.2f}s")
    
    benchmark.save_results()
    print("\nBenchmark complete!")
